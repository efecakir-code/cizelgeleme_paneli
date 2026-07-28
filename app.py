import streamlit as st
import pandas as pd
import json
import os
import subprocess
import time
import datetime
import plotly.express as px
import plotly.graph_objects as go
import uuid
import copy
import tempfile
from pathlib import Path

# Excel parser import (varsayılan yolları da import edebiliriz)
try:
    from build_143_input import build_payload, DEFAULT_ROUTE_INPUT, DEFAULT_PRIORITY_WORKBOOK
except ImportError:
    st.error("build_143_input.py dosyası bulunamadı. Lütfen aynı dizinde olduğundan emin olun.")

st.set_page_config(page_title="Üretim Planlama Çizelgeleme", layout="wide")

# Sabit dosya ve klasör yolları
BASE_DATA = "master_net_manufacturing_input.json"
CUSTOM_INPUT = "custom_input.json"
OUTPUT_DIR = "custom_output"
FIFO_OUTPUT_DIR = "fifo_output"
CHECKPOINT_FILE = os.path.join(OUTPUT_DIR, "checkpoint.json")
FIFO_CHECKPOINT_FILE = os.path.join(FIFO_OUTPUT_DIR, "FIFO_KANIT.json")

if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)
if not os.path.exists(FIFO_OUTPUT_DIR):
    os.makedirs(FIFO_OUTPUT_DIR, exist_ok=True)

@st.cache_data
def load_base_data(uploaded_file=None):
    if uploaded_file is not None:
        # Geçici olarak Excel dosyasını kaydet
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getvalue())
            tmp_path = tmp.name
        try:
            # build_payload ile json verisini direkt RAM'de üret (diske yazmadan)
            data, audit = build_payload(
                workbook_path=Path(tmp_path), 
                route_input=DEFAULT_ROUTE_INPUT,
                priority_workbook=DEFAULT_PRIORITY_WORKBOOK
            )
            os.remove(tmp_path)
            st.success("Excel başarıyla okundu ve json formatına dönüştürüldü.")
        except Exception as e:
            st.error(f"Excel okunurken hata oluştu: {e}")
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
            return None, pd.DataFrame(), {}
    else:
        if not os.path.exists(BASE_DATA):
            st.warning(f"Sistemde `{BASE_DATA}` bulunamadı, lütfen Excel dosyası yükleyin.")
            return None, pd.DataFrame(), {}
        with open(BASE_DATA, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
    units = data.get("units", [])
    
    # Alternatör Şablonlarını Çıkar
    templates = {}
    for u in units:
        code = u.get("alternator_code")
        if code and code not in templates:
            templates[code] = copy.deepcopy(u)
            
    df = pd.DataFrame(units)
    # Görüntülenecek sütunları seç
    if not df.empty:
        display_cols = ["unit_id", "order_job_id", "alternator_code", "model", "priority", "planned_ship_date", "ready_date", "project"]
        # Eğer bu alanlar excelden çekilmiş veride (bazılarında) yoksa hata almamak için filtreleyelim
        display_cols = [c for c in display_cols if c in df.columns]
        df_display = df[display_cols].copy()
    else:
        df_display = pd.DataFrame()
        
    return data, df_display, templates

def run_solver(time_limit_sec, cp_round, ga_stage, full_cp, pop_size, elite_count, crossover_rate, mutation_rate, stagnation_gen, seed_val, vns_interval=10, vns_elite=3, diversify_rate=0.35, only_cp=False):
    """Arka planda CP-SAT & GA'yı çalıştırır, logları canlı gösterir"""
    import sys
    
    cmd = [
        sys.executable, "HGA_VNS_CP_SAT_MAIN.py",
        "--input", CUSTOM_INPUT,
        "--output-dir", OUTPUT_DIR,
        "--population-size", str(pop_size),
        "--elite-count", str(elite_count),
        "--crossover-rate", str(crossover_rate),
        "--mutation-rate", str(mutation_rate),
        "--stagnation-generations", str(stagnation_gen),
        "--vns-interval-generations", str(vns_interval),
        "--vns-elite-count", str(vns_elite),
        "--diversify-rate", str(diversify_rate),
        "--seed", str(seed_val),
        "--checkpoint-min-seconds", "5",
        "--skip-excel"
    ]
    
    import os
    fifo_state = os.path.join(FIFO_OUTPUT_DIR, "best_solution_state.json")
    if os.path.exists(fifo_state):
        cmd.extend(["--initial-state", fifo_state])
    
    if only_cp:
        cmd.extend([
            "--stage-seconds", str(time_limit_sec),
            "--cp-seconds-per-round", str(time_limit_sec),
            "--ga-seconds-per-stage", "0",
            "--full-cp-stagnation-seconds", "0",
            "--full-cp-seconds", str(full_cp)
        ])
    else:
        cmd.extend([
            "--stage-seconds", str(time_limit_sec),
            "--cp-seconds-per-round", str(cp_round),
            "--ga-seconds-per-stage", str(ga_stage),
            "--full-cp-seconds", str(full_cp)
        ])
    
    # Eski checkpoint'i temizle
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)
        
    st.info("Çözücü başlatıldı, lütfen bekleyin...")
    
    timer_container = st.empty()
    log_container = st.empty()
    log_lines = []
    
    start_time = time.time()
    total_alg_time = (time_limit_sec * 3) + full_cp
    is_optimal_run = time_limit_sec > 10000  # Limitsiz run check
    
    try:
        # bufsize=1 for line buffering, stderr=subprocess.STDOUT merges errors into stdout
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)
        
        last_update_time = time.time()
        
        while True:
            line = process.stdout.readline()
            if line:
                log_lines.append(line.strip())
                
                # Sadece belli aralıklarla UI'ı güncelle (Arayüzün çökmesini/kilitlenmesini engellemek için)
                current_time = time.time()
                if current_time - last_update_time > 0.25:
                    display_text = "\n".join(log_lines[-20:])
                    log_container.code(display_text, language="bash")
                    
                    elapsed = current_time - start_time
                    if is_optimal_run:
                        timer_container.info(f"⏳ **Limitsiz Optimal Çözüm Arıyor** | Geçen Süre: {elapsed:.0f} sn")
                    else:
                        remaining = max(0, total_alg_time - elapsed)
                        timer_container.info(f"⏳ **Maksimum Kalan Süre:** {remaining:.0f} saniye")
                        
                    last_update_time = current_time
                
            retcode = process.poll()
            if not line and retcode is not None:
                # İşlem bittiğinde son kalan logları da kesin olarak ekrana bas
                display_text = "\n".join(log_lines[-20:])
                log_container.code(display_text, language="bash")
                break
            
            # Zaman sınırını kontrol et (eğer subprocess takılırsa diye safety timeout)
            max_expected_time = (time_limit_sec * 3) + full_cp + 120
            if time.time() - start_time > max_expected_time:
                process.terminate()
                st.warning(f"Çözücü güvenlik süresini ({max_expected_time} sn) aştığı için durdurulmaya çalışılıyor...")
                time.sleep(2)
                break
                
        if os.path.exists(CHECKPOINT_FILE):
            st.success(f"Çözüm başarıyla tamamlandı! ({time.time()-start_time:.1f} saniye)")
            return True
        else:
            st.error("Çözüm tamamlandı fakat checkpoint.json dosyası bulunamadı. Lütfen yukarıdaki terminal loglarını inceleyin.")
            return False
            
    except Exception as e:
        st.error(f"Çözücü çalıştırılırken hata oluştu: {e}")
        return False

def draw_gantt(out_dir=OUTPUT_DIR, is_fifo=False):
    chk_file = os.path.join(out_dir, "checkpoint.json" if not is_fifo else "FIFO_KANIT.json")
    if not os.path.exists(chk_file):
        return
        
    with open(chk_file, 'r', encoding='utf-8') as f:
        chk = json.load(f)
        
    schedule_file = os.path.join(out_dir, "schedule.csv")
    if not os.path.exists(schedule_file):
        st.warning("Gantt çizdirilecek veri bulunamadı (schedule.csv yok).")
        return
        
    df = pd.read_csv(schedule_file)
    if df.empty:
        st.warning("Gantt çizdirilecek veri bulunamadı (tablo boş).")
        return
    
    # Zamanları gerçek saate çevirme
    base_date_str = chk.get("start_time", "2026-07-21T00:00:00")
    if "T" not in base_date_str:
        base_date_str += "T00:00:00"
    try:
        base_dt = datetime.datetime.fromisoformat(base_date_str.split(".")[0])
    except:
        base_dt = datetime.datetime(2026, 7, 21)
        
    # start_min ve end_min zaten dakika cinsinden olduğu için ekstra bölme yapmıyoruz
    def mins_to_dt(val):
        return base_dt + datetime.timedelta(minutes=float(val))
        
    # start_min ve end_min kullanıyoruz (schedule.csv formatı)
    df['Start'] = df['start_min'].apply(mins_to_dt)
    df['End'] = df['end_min'].apply(mins_to_dt)
    
    # Calculate Net Makespan (Excluding Gaps)
    intervals = []
    for _, row in df.iterrows():
        intervals.append((row["start_min"], row["end_min"]))
    intervals.sort(key=lambda x: x[0])
    
    merged_intervals = []
    for interval in intervals:
        if not merged_intervals:
            merged_intervals.append(interval)
        else:
            prev = merged_intervals[-1]
            if interval[0] <= prev[1]:
                merged_intervals[-1] = (prev[0], max(prev[1], interval[1]))
            else:
                merged_intervals.append(interval)
                
    net_makespan = sum(i[1] - i[0] for i in merged_intervals)
    gross_makespan = chk.get('objectives', {}).get('makespan_min', 0)

    fig = px.timeline(
        df, 
        x_start="Start", 
        x_end="End", 
        y="machine", 
        color="unit_id", 
        hover_data=["operation", "start_min", "end_min", "sequence"],
        title=f"Üretim Çizelgesi - Net (Aktif) Makespan: {net_makespan:.1f} dk | Brüt: {gross_makespan:.1f} dk",
        height=700
    )
    fig.update_yaxes(autorange="reversed")
    fig.update_layout(
        showlegend=False,
        xaxis=dict(
            rangeslider=dict(
                visible=True
            ),
            type="date"
        )
    )
    st.plotly_chart(fig, use_container_width=True)
    
    return {"chk": chk, "net_makespan": net_makespan}

def check_password():
    """Returns `True` if the user had the correct password."""

    def password_entered():
        """Checks whether a password entered by the user is correct."""
        if st.session_state["username"] == "admin" and st.session_state["password"] == "admin2024":
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # don't store password
            del st.session_state["username"]
        else:
            st.session_state["password_correct"] = False

    if st.session_state.get("password_correct", False):
        return True

    # Show input for username and password.
    st.markdown("## 🔒 Üretim Planlama Çizelgeleme Paneli Giriş")
    st.text_input("Kullanıcı Adı", key="username")
    st.text_input("Şifre", type="password", key="password")
    if st.button("Giriş Yap", on_click=password_entered):
        pass
        
    if "password_correct" in st.session_state and not st.session_state["password_correct"]:
        st.error("Kullanıcı adı veya şifre hatalı!")
        
    return False

def main():
    if not check_password():
        return
        
    st.title("Üretim Planlama Çizelgeleme Paneli")
    
    # Yeni eklenti: Excel yükleme opsiyonu
    with st.expander("Excel'den Veri Yükle (Opsiyonel)", expanded=False):
        st.write("Eğer sistemde var olan JSON'ı kullanmak yerine güncel Excel'den çekim yapmak isterseniz:")
        uploaded_file = st.file_uploader("Net_Ihtiyac_Detay ve 157_Tekil_Is sekmeleri olan güncel kaynak Excel'i seçin", type=["xlsx"])
    
    data, df_base, templates = load_base_data(uploaded_file)
    if data is None:
        return
        
    st.header("1. Mevcut Siparişlerden Seçim Yapın")
    st.write(f"Şu anki verisetinde toplam **{len(df_base)}** adet kayıtlı sipariş bulunmaktadır.")
    
    if "select_all_toggle" not in st.session_state:
        st.session_state.select_all_toggle = True
    if "editor_key" not in st.session_state:
        st.session_state.editor_key = 0

    col_s1, col_s2, _ = st.columns([2, 2, 8])
    with col_s1:
        if st.button("✅ Tümünü Seç"):
            st.session_state.select_all_toggle = True
            st.session_state.editor_key += 1
            st.rerun()
    with col_s2:
        if st.button("❌ Seçimi Kaldır"):
            st.session_state.select_all_toggle = False
            st.session_state.editor_key += 1
            if "added_units" in st.session_state:
                st.session_state.added_units = []
            st.rerun()
            
    df_selectable = df_base.copy()
    df_selectable.insert(0, "Seç", st.session_state.select_all_toggle)
    
    edited_df = st.data_editor(
        df_selectable,
        key=f"data_editor_{st.session_state.editor_key}",
        hide_index=True,
        use_container_width=True,
        column_config={
            "Seç": st.column_config.CheckboxColumn("Çizelgeye Al", default=True)
        }
    )
    
    selected_unit_ids = edited_df[edited_df["Seç"] == True]["unit_id"].tolist() if "Seç" in edited_df.columns else []
    
    st.header("2. Yeni Sipariş Ekle")
    with st.expander("Yeni Sipariş Formu", expanded=False):
        st.info("Aşağıdaki alanlar tablodaki başlıklarla birebir eşleşmektedir.")
        col1, col2, col3 = st.columns(3)
        with col1:
            new_unit_id = st.text_input("unit_id (Otomatik atanır, değiştirebilirsiniz)", value=f"NEW-{uuid.uuid4().hex[:4].upper()}")
            new_code = st.selectbox("alternator_code", list(templates.keys()))
            new_qty = st.number_input("Adet (Birden fazlaysa ID sonuna sayı eklenir)", min_value=1, value=1, step=1)
            
        with col2:
            new_order_id = st.text_input("order_job_id (Canias Sipariş No)", value="YENI-SIPARIS")
            new_ready = st.date_input("ready_date (Üretime Giriş Tarihi)", datetime.date.today())
            new_due = st.date_input("planned_ship_date (Teslim Tarihi)", datetime.date.today() + datetime.timedelta(days=30))
            
        with col3:
            new_model = st.text_input("model (Stok kodundan otomatik gelir)", value=templates.get(new_code, {}).get("model", ""), disabled=True)
            new_priority = st.selectbox("priority (Öncelik)", ["Normal", "Yüksek", "Acil"])
            new_project = st.text_input("project (Müşteri / Proje Adı)", value="Yeni Proje")
            
        if st.button("Siparişleri Listeye Ekle"):
            if "added_units" not in st.session_state:
                st.session_state.added_units = []
                
            base_template = templates[new_code]
            
            for i in range(new_qty):
                new_unit = copy.deepcopy(base_template)
                
                # Assign values exactly
                current_unit_id = f"{new_unit_id}-{i+1}" if new_qty > 1 else new_unit_id
                new_unit["unit_id"] = current_unit_id
                new_unit["order_job_id"] = f"{new_order_id}-{i+1}" if new_qty > 1 else new_order_id
                new_unit["priority"] = new_priority
                weights = data.get("parameters", {}).get("priority_weights", {"Acil":3, "Yüksek":2, "Normal":1})
                new_unit["priority_weight"] = weights.get(new_priority, 1)
                
                new_unit["ready_date"] = new_ready.isoformat()
                new_unit["planned_ship_date"] = new_due.isoformat()
                new_unit["model_due_date"] = new_due.isoformat()
                new_unit["project"] = new_project
                
                base_date = datetime.datetime.fromisoformat(data["base_date"])
                
                def calc_minutes(d):
                    diff = d - base_date.date()
                    days = max(0, diff.days)
                    workdays = int(days * (5/7))
                    return workdays * data.get("daily_work_minutes", 540) * data.get("time_scale", 100)
                    
                new_unit["release"] = calc_minutes(new_ready)
                new_unit["due"] = calc_minutes(new_due) + (data.get("daily_work_minutes", 540) * data.get("time_scale", 100))
                
                st.session_state.added_units.append(new_unit)
            st.success(f"{new_qty} adet sipariş listeye eklendi.")

    if "added_units" in st.session_state and st.session_state.added_units:
        st.write("### Yeni Eklenen Siparişler")
        st.dataframe(pd.DataFrame(st.session_state.added_units)[["unit_id", "order_job_id", "alternator_code", "priority", "planned_ship_date", "project"]])
        
        if st.button("Eklenenleri Temizle"):
            st.session_state.added_units = []
            st.rerun()

    st.header("3. Temel Durum (FIFO Baseline)")
    st.write("Optimizasyona başlamadan önce, fabrikanın mevcut standart mantığıyla (İlk Giren İlk Çıkar) nasıl bir çizelge oluşacağını görebilirsiniz.")
    
    if st.button("📊 Seçili Siparişler İçin FIFO Başlangıç Çizelgesini Oluştur ve İncele", type="primary", use_container_width=True):
        custom_data = copy.deepcopy(data)
        final_units = [u for u in data["units"] if u["unit_id"] in selected_unit_ids]
        if "added_units" in st.session_state:
            final_units.extend(st.session_state.added_units)
            
        custom_data["units"] = final_units
        custom_data["unit_count"] = len(final_units)
        
        with open(CUSTOM_INPUT, 'w', encoding='utf-8') as f:
            json.dump(custom_data, f, ensure_ascii=False, indent=2)
            
        import subprocess
        import sys
        st.info("FIFO çizelgesi oluşturuluyor...")
        # FIFO betiğini çalıştır
        subprocess.run([sys.executable, "FIFO_BASLANGIC_OLUSTUR.py", "--input", CUSTOM_INPUT, "--output-dir", FIFO_OUTPUT_DIR], check=True)
        
        # Sonucu yükle
        fifo_chk_path = os.path.join(FIFO_OUTPUT_DIR, "FIFO_KANIT.json")
        if os.path.exists(fifo_chk_path):
            with open(fifo_chk_path, 'r', encoding='utf-8') as f:
                st.session_state.fifo_baseline_obj = json.load(f)
            st.session_state.show_fifo = True
            st.success("FIFO Referans Çizelgesi Hazır!")
        else:
            st.error("FIFO Çizelgesi oluşturulamadı!")
            
    if st.session_state.get("show_fifo", False):
        st.subheader("İlk Durum (Optimizasyonsuz) Sonuçları")
        fifo_obj = st.session_state.fifo_baseline_obj
        
        metrics_container = st.container()
        
        fifo_gantt_result = draw_gantt(out_dir=FIFO_OUTPUT_DIR, is_fifo=True)
        if fifo_gantt_result:
            st.session_state.fifo_chk = fifo_gantt_result
            
        fifo_net_makespan = st.session_state.fifo_chk.get("net_makespan", 0) if "fifo_chk" in st.session_state else 0
        
        with metrics_container:
            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            col_f1.metric("FIFO Brüt Makespan", f"{fifo_obj.get('makespan_min',0):.1f} dk")
            col_f2.metric("FIFO Net Makespan", f"{fifo_net_makespan:.1f} dk")
            col_f3.metric("FIFO Toplam Gecikme", f"{fifo_obj.get('weighted_tardiness',0):.1f} dk")
            col_f4.metric("FIFO Toplam Setup", f"{fifo_obj.get('total_setup_min',0):.1f} dk")

    st.header("4. Çizelgeleme ve Gelişmiş Ayarlar (Optimizasyon)")
    # Detaylı Süre/Algoritma Ayarları
    with st.expander("Gelişmiş Algoritma Süre ve Parametre Ayarları", expanded=True):
        st.markdown("Algoritmanın çalışma sürelerini ve derinliğini buradan ince ayar ile belirleyebilirsiniz. Süreleri artırmak daha iyi (kısa makespan) sonuç bulma ihtimalini artırır, ancak bekleme süresini uzatır.")
        col_t1, col_t2 = st.columns(2)
        with col_t1:
            time_limit = st.slider("Her Aşama İçin Maks Süre (--stage-seconds)", min_value=15, max_value=600, value=60, step=15, help="Optimizasyonun her bir aşaması (toplam 3 aşama var) için ayrılan maksimum saniye.\n\n📈 Artırılırsa: Daha uzun süre daha iyi sonuçlar aranır ancak çözüm gecikir.\n\n📉 Azaltılırsa: Çözüm hızlı biter ancak sonuçlar kalitesiz olabilir.")
            ga_max = max(10, time_limit - 1)
            ga_stage = st.slider("GA Çözüm Süresi (--ga-seconds-per-stage)", min_value=5, max_value=ga_max, value=min(20, ga_max), step=5, help="Bu sürenin ne kadarının Genetik Algoritmaya ayrılacağı. Kalanı CP-SAT'a verilir.\n\n📈 Artırılırsa: Global arama alanı daha fazla taranır, genetik çeşitlilik artar.\n\n📉 Azaltılırsa: CP-SAT (yerel matematiksel arama) daha çok zaman bulur.")
            pop_size = st.number_input("Popülasyon Büyüklüğü (Population Size)", min_value=10, max_value=1000, value=64, step=10, help="Genetik Algoritmadaki her nesildeki birey (farklı çizelge alternatifi) sayısı.\n\n📈 Artırılırsa: Çok daha geniş bir arama havuzu oluşur, daha iyi çözümler bulunabilir ancak algoritma yavaşlar ve bellek kullanımı artar.\n\n📉 Azaltılırsa: Hızlı çalışır ama erken durgunluğa girme riski doğar.")
        with col_t2:
            cp_round = st.slider("Ara CP-SAT Süresi (--cp-seconds-per-round)", min_value=5, max_value=60, value=15, step=5, help="Hibrit sistemde CP-SAT'ın ara iyileştirmeleri için kullanacağı saniye.\n\n📈 Artırılırsa: Ara adımlarda çok güçlü matematiksel optimizasyon yapılır, ancak genel tur süresi uzar.")
            full_cp = st.slider("Final CP-SAT Süresi (--full-cp-seconds)", min_value=5, max_value=300, value=20, step=5, help="Tüm aşamalar bittikten sonra en son uygulanan yoğun cila (iyileştirme) süresi.\n\n📈 Artırılırsa: Tüm işlemlerin son noktasında çok daha iyi yerel çözümler için uzun süre harcanır.")
            pure_cp_time = st.slider("Sadece CP-SAT İçin Süre (sn)", min_value=15, max_value=600, value=60, step=15, help="Sadece CP-SAT modunda veya Limitsiz seçeneğinde algoritmanın kullanacağı temel tur süresi limiti.")
            
        st.markdown("### 🧬 HGA Parametreleri (Gelişmiş)")
        col_hga1, col_hga2, col_hga3 = st.columns(3)
        with col_hga1:
            seed_val = st.number_input("Rastgelelik Tohumu (Random Seed)", value=42, help="Sonuçların tekrarlanabilirliğini sağlar. Farklı bir tohum farklı başlangıç genetik dizilimleri oluşturur. Belirli bir düzende kalması için genelde sabit bırakılır.")
            crossover_rate = st.slider("Çaprazlama Oranı (Crossover Rate)", min_value=0.0, max_value=1.0, value=0.85, step=0.01, help="İki iyi çizelgeden (ebeveyn) yeni bir çizelge (çocuk) oluşturma olasılığı.\n\n📈 Artırılırsa: Çözüm uzayında daha agresif sıçramalar ve birleşmeler olur, iyileşme hızlanabilir.\n\n📉 Azaltılırsa: Çözümler daha çok orijinal hallerinde kalır, mutasyona bağımlı hale gelir.")
            diversify_rate = st.slider("Çeşitlendirme Oranı (Diversification)", min_value=0.05, max_value=0.90, value=0.35, step=0.05, help="Durgunluk (stagnation) anında popülasyonun yüzde kaçının rastgele yepyeni bireylerle (kan değişimi) değiştirileceğini belirler.\n\n📈 Artırılırsa: Havuza çok fazla yeni gen girer, tıkanıklık aşılabilir ama mevcut iyi yapılar bozulabilir.\n\n📉 Azaltılırsa: Havuz korunur ama yerel optimumdan çıkmak zorlaşır.")
        with col_hga2:
            elite_count = st.number_input("Elit Birey Sayısı (Elite Size)", min_value=1, max_value=50, value=8, help="Bir nesilden diğerine doğrudan aktarılan, dokunulmayan (bozulmayan) en iyi birey sayısı.\n\n📈 Artırılırsa: Bulunan iyi çözümler kesinlikle korunur, ancak çeşitlilik azalabilir.\n\n📉 Azaltılırsa: Çeşitlilik artar fakat iyi çizelgelerin kaybolma riski doğar.")
            mutation_rate = st.slider("Mutasyon Oranı (Mutation Rate)", min_value=0.0, max_value=1.0, value=0.10, step=0.01, help="Bir çizelgenin yapısının rastgele bir mutasyonla (örn: bir işin makinesini değiştirerek) bozulma/değişme olasılığı.\n\n📈 Artırılırsa: Yeni keşifler (diversification) çok artar, durgunluk bozulur ama var olan iyi yapılar yıkılabilir.\n\n📉 Azaltılırsa: Sistemin dengesi korunur ama yerel optimumlara (lokal sıkışma) saplanma riski artar.")
        with col_hga3:
            stagnation_gen = st.number_input("Durgunluk Ölçütü (Stagnation)", min_value=1, max_value=200, value=25, help="Kaç nesil boyunca iyileşme olmazsa sistemin 'durgunluk' (stagnation) moduna geçip çeşitlendirme mekanizmalarını tetikleyeceğini belirler.\n\n📈 Artırılırsa: Sabırla aynı havuz taranmaya devam eder.\n\n📉 Azaltılırsa: Sistem sabırsız davranıp çok çabuk mutasyon ve çeşitlendirmelere başvurur.")
            vns_interval = st.number_input("VNS Uygulanma Sıklığı", min_value=1, max_value=50, value=10, help="Kaç nesilde bir elit bireylere VNS (Değişken Komşuluk Araması) uygulanacağını belirler.\n\n📈 Artırılırsa: VNS daha az yapılır, GA daha hızlı çalışır.\n\n📉 Azaltılırsa: Daha sık VNS yapılır, kaliteli sonuç ihtimali artar ancak süre uzayabilir.")
            vns_elite = st.number_input("VNS Elit Birey Sayısı", min_value=1, max_value=20, value=3, help="VNS'nin uygulanacağı en iyi (elit) birey sayısını belirler.")
            st.info("💡 Çeşitlendirme (Diversification), POX ve Uniform Crossover yöntemleri motor seviyesinde aktiftir.")
    
    col_b1, col_b2, col_b3 = st.columns(3)
    with col_b1:
        run_hybrid = st.button("🚀 Hibrit Çizelgele (HGA + CP-SAT)", type="primary", use_container_width=True)
    with col_b2:
        run_pure_cp = st.button("⚙️ Sadece CP-SAT ile Çöz", type="secondary", use_container_width=True)
    with col_b3:
        run_optimal = st.button("🧠 Optimal Bulana Kadar Çöz", type="secondary", use_container_width=True)
        
    if run_hybrid or run_pure_cp or run_optimal:
        st.session_state.run_hybrid = run_hybrid
        st.session_state.run_pure_cp = run_pure_cp
        st.session_state.run_optimal = run_optimal
        st.session_state.solver_started = True
        st.session_state.solver_finished = False

    if st.session_state.get("solver_started", False):
        run_hybrid_state = st.session_state.get("run_hybrid", False)
        run_pure_cp_state = st.session_state.get("run_pure_cp", False)
        run_optimal_state = st.session_state.get("run_optimal", False)
        
        custom_data = copy.deepcopy(data)
        final_units = [u for u in data["units"] if u["unit_id"] in selected_unit_ids]
        if "added_units" in st.session_state:
            final_units.extend(st.session_state.added_units)
            
        custom_data["units"] = final_units
        custom_data["unit_count"] = len(final_units)
        
        with open(CUSTOM_INPUT, 'w', encoding='utf-8') as f:
            json.dump(custom_data, f, ensure_ascii=False, indent=2)
            
        is_pure = run_pure_cp_state or run_optimal_state
        
        if run_optimal_state:
            time_limit_val = 86400  # 24 saat
            cp_round_val = 86400
            ga_stage_val = 0
            full_cp_val = 86400
            islem_tipi = "Limitsiz Optimal CP-SAT (Çok Uzun Sürebilir)"
        else:
            time_limit_val = pure_cp_time if run_pure_cp_state else time_limit
            cp_round_val = cp_round
            ga_stage_val = ga_stage
            full_cp_val = full_cp
            islem_tipi = "Sadece CP-SAT" if run_pure_cp_state else "Hibrit (HGA+CP-SAT)"
            
        st.write(f"**Toplam {len(final_units)} iş {islem_tipi} ile çizelgeleniyor...**")
        
        if not st.session_state.get("solver_finished", False):
            success = run_solver(time_limit_val, cp_round_val, ga_stage_val, full_cp_val, pop_size, elite_count, crossover_rate, mutation_rate, stagnation_gen, seed_val, vns_interval, vns_elite, only_cp=is_pure)
            if success:
                st.session_state.solver_finished = True
            else:
                st.session_state.solver_started = False
                st.stop()
                
        if st.session_state.get("solver_finished", False):
            gantt_result = draw_gantt()
            if gantt_result:
                chk = gantt_result["chk"]
                opt_net_makespan = gantt_result["net_makespan"]
                st.success("İşlem Tamamlandı!")
                
                objectives = chk.get("objectives", {})
                opt_gross = objectives.get("makespan_min", 0)
                opt_tard = objectives.get("weighted_tardiness", 0)
                opt_setup = objectives.get("total_setup_min", 0)
                
                fifo_gross, fifo_net, fifo_tard, fifo_setup = 0, 0, 0, 0
                imp_gross, imp_net, imp_tard, imp_setup = 0, 0, 0, 0
                
                # İyileşme (İyileştirme) Hesaplaması ve Gösterimi
                if "fifo_baseline_obj" in st.session_state:
                    st.subheader("Optimizasyon İyileşme Oranları")
                    fifo_obj = st.session_state.fifo_baseline_obj
                    fifo_gross = fifo_obj.get("makespan_min", 0)
                    fifo_tard = fifo_obj.get("weighted_tardiness", 0)
                    fifo_setup = fifo_obj.get("total_setup_min", 0)
                    
                    # Net makespan from fifo
                    if "fifo_chk" in st.session_state and st.session_state.fifo_chk:
                        fifo_net = st.session_state.fifo_chk.get("net_makespan", 0)
                        
                    def calc_imp(f, o):
                        if f == 0: return 0
                        return ((f - o) / f) * 100
                        
                    imp_gross = calc_imp(fifo_gross, opt_gross)
                    imp_net = calc_imp(fifo_net, opt_net_makespan) if fifo_net > 0 else 0
                    imp_tard = calc_imp(fifo_tard, opt_tard)
                    imp_setup = calc_imp(fifo_setup, opt_setup)
                    
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Brüt Makespan İyileşmesi", f"%{imp_gross:.1f}", delta=f"{fifo_gross-opt_gross:.1f} dk", delta_color="normal")
                    c2.metric("Net Makespan İyileşmesi", f"%{imp_net:.1f}", delta=f"{fifo_net-opt_net_makespan:.1f} dk", delta_color="normal")
                    c3.metric("Gecikme İyileşmesi", f"%{imp_tard:.1f}", delta=f"{fifo_tard-opt_tard:.1f} dk", delta_color="normal")
                    c4.metric("Setup İyileşmesi", f"%{imp_setup:.1f}", delta=f"{fifo_setup-opt_setup:.1f} dk", delta_color="normal")
                
                st.subheader("Sonuçları İndir")
        
                try:
                    excel_path = os.path.join(OUTPUT_DIR, "cizelge_sonuclari.xlsx")
                    
                    # 1. Gantt Verisi (schedule.csv)
                    schedule_file = os.path.join(OUTPUT_DIR, "schedule.csv")
                    df_schedule = pd.read_csv(schedule_file)
                    for col in df_schedule.select_dtypes(include=['datetimetz']).columns:
                        df_schedule[col] = df_schedule[col].dt.tz_localize(None)
                        
                    # 2. Makine Kullanım Verisi (machine_utilization.csv)
                    machine_file = os.path.join(OUTPUT_DIR, "machine_utilization.csv")
                    if os.path.exists(machine_file):
                        df_machine = pd.read_csv(machine_file)
                    else:
                        df_machine = pd.DataFrame({"Hata": ["Makine verisi bulunamadı."]})
                        
                    # 3. Amaç Fonksiyonları (checkpoint.json)
                    df_obj = pd.DataFrame([{
                        "FIFO Brüt Makespan (dk)": fifo_gross,
                        "Opt. Brüt Makespan (dk)": opt_gross,
                        "Brüt İyileşme (%)": imp_gross,
                        
                        "FIFO Net Makespan (dk)": fifo_net,
                        "Opt. Net Makespan (dk)": opt_net_makespan,
                        "Net İyileşme (%)": imp_net,
                        
                        "FIFO Toplam Gecikme (dk)": fifo_tard,
                        "Opt. Toplam Gecikme (dk)": opt_tard,
                        "Gecikme İyileşme (%)": imp_tard,
                        
                        "FIFO Toplam Setup (dk)": fifo_setup,
                        "Opt. Toplam Setup (dk)": opt_setup,
                        "Setup İyileşme (%)": imp_setup
                    }])
                    
                    # Excel'e birden fazla sayfa (sheet) olarak yazdır
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        df_obj.to_excel(writer, sheet_name="Amaç Fonksiyonları", index=False)
                        df_machine.to_excel(writer, sheet_name="Makine Kullanımı", index=False)
                        df_schedule.to_excel(writer, sheet_name="Çizelge (Liste)", index=False)
                        
                        # 4. Görsel Gantt Şeması Ekleme (Optimizasyon)
                        workbook = writer.book
                        ws = workbook.create_sheet("Opt. Gantt Görseli")
                        
                        from openpyxl.styles import PatternFill
                        from openpyxl.utils import get_column_letter
                        import hashlib
                        
                        def get_color(uid):
                            hash_obj = hashlib.md5(str(uid).encode())
                            hex_col = hash_obj.hexdigest()[:6]
                            return "FF" + hex_col
                            
                        def draw_excel_gantt(worksheet, schedule_df, makespan_val):
                            factor = 30 # Her sütun 30 dakika
                            makespan_slots = int(makespan_val / factor) + 2
                            worksheet.cell(row=1, column=1, value="Makine \\ Saat")
                            worksheet.column_dimensions['A'].width = 20
                            
                            for s in range(makespan_slots):
                                col_letter = get_column_letter(s + 2)
                                hour_label = f"{int((s*factor)/60):02d}:{(s*factor)%60:02d}"
                                worksheet.cell(row=1, column=s+2, value=hour_label)
                                worksheet.column_dimensions[col_letter].width = 6
                                
                            macs = sorted(schedule_df['machine'].unique())
                            r_idx = 2
                            
                            setup_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid") # Gri setup
                            
                            for m in macs:
                                worksheet.cell(row=r_idx, column=1, value=m)
                                m_tasks = schedule_df[schedule_df['machine'] == m]
                                for _, task in m_tasks.iterrows():
                                    s_min = task['start_min']
                                    e_min = task['end_min']
                                    setup_s_min = task.get('setup_start_min', s_min)
                                    
                                    setup_s_col = int(setup_s_min / factor) + 2
                                    s_col = int(s_min / factor) + 2
                                    e_col = int(e_min / factor) + 2
                                    
                                    uid = str(task['unit_id']).replace("MOCK-", "").replace("NEW-", "")[:8]
                                    op = task.get("operation", "")
                                    
                                    # Setup boyama
                                    if setup_s_col < s_col:
                                        for c_idx in range(setup_s_col, s_col):
                                            cell = worksheet.cell(row=r_idx, column=c_idx)
                                            cell.fill = setup_fill
                                            if c_idx == setup_s_col:
                                                cell.value = f"Ayar:{int(s_min-setup_s_min)}d"
                                                
                                    # İşlem boyama
                                    color = get_color(task['unit_id'])
                                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                    # En az 1 hücre boyansın
                                    if s_col == e_col: e_col = s_col + 1
                                    
                                    for c_idx in range(s_col, e_col):
                                        cell = worksheet.cell(row=r_idx, column=c_idx)
                                        cell.fill = fill
                                        if c_idx == s_col:
                                            cell.value = f"{uid} {op} ({int(s_min)}-{int(e_min)})"
                                            
                                r_idx += 1
                                
                        draw_excel_gantt(ws, df_schedule, opt_gross)
                        
                        # 5. FIFO Çizelge ve Gantt Görseli (Eğer varsa)
                        fifo_schedule_file = os.path.join(FIFO_OUTPUT_DIR, "schedule.csv")
                        if os.path.exists(fifo_schedule_file):
                            df_fifo = pd.read_csv(fifo_schedule_file)
                            for col in df_fifo.select_dtypes(include=['datetimetz']).columns:
                                df_fifo[col] = df_fifo[col].dt.tz_localize(None)
                            df_fifo.to_excel(writer, sheet_name="FIFO Çizelge (Liste)", index=False)
                            
                            ws_fifo = workbook.create_sheet("FIFO Gantt Görseli")
                            draw_excel_gantt(ws_fifo, df_fifo, fifo_gross)

                    
                    with open(excel_path, "rb") as f:
                        st.download_button(
                            label="📥 Kapsamlı Excel Raporunu İndir",
                            data=f,
                            file_name="cizelgeleme_kapsamli_rapor.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.error(f"Excel'e aktarılırken hata oluştu: {str(e)}")
                    st.info("İpucu: 'openpyxl' kütüphanesi eksik olabilir. Lütfen terminalde 'pip install openpyxl' çalıştırın.")

if __name__ == "__main__":
    main()
