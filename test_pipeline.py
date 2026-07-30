import json, sys, subprocess, os, time, pandas as pd
import datetime, copy, hashlib
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 1. Create a minimal custom_input.json from master
master = json.load(open("master_net_manufacturing_input.json", "r"))
master["units"] = master["units"][:3]
master["unit_count"] = 3
json.dump(master, open("custom_input.json", "w"), indent=2)

print("Running FIFO...")
res1 = subprocess.run([sys.executable, "FIFO_BASLANGIC_OLUSTUR.py", "--input", "custom_input.json", "--output-dir", "fifo_output"])
print("FIFO Retcode:", res1.returncode)

print("Running CP-SAT...")
cmd = [
    sys.executable, "HGA_VNS_CP_SAT_MAIN.py",
    "--input", "custom_input.json",
    "--output-dir", "custom_output",
    "--population-size", "20",
    "--checkpoint-min-seconds", "5",
    "--skip-excel",
    "--stage-seconds", "15",
    "--cp-seconds-per-round", "15",
    "--ga-seconds-per-stage", "0",
    "--full-cp-seconds", "15"
]
res2 = subprocess.run(cmd)
print("CP-SAT Retcode:", res2.returncode)

# Now test Excel Generation
print("Testing Excel generation...")
fifo_chk = json.load(open("fifo_output/FIFO_KANIT.json"))
fifo_gross = fifo_chk.get("makespan_min", 0)

opt_chk = json.load(open("custom_output/checkpoint.json"))
opt_gross = opt_chk.get("objectives", {}).get("makespan_min", 0)

df_schedule = pd.read_csv("custom_output/schedule.csv")
for col in df_schedule.select_dtypes(include=['datetimetz']).columns:
    df_schedule[col] = df_schedule[col].dt.tz_localize(None)

excel_path = "test_cizelge.xlsx"
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_schedule.to_excel(writer, sheet_name="Çizelge (Liste)", index=False)
    
    workbook = writer.book
    ws = workbook.create_sheet("Opt. Gantt Görseli")
    
    def get_color(uid):
        return "FF" + hashlib.md5(str(uid).encode()).hexdigest()[:6]
        
    def draw_excel_gantt(worksheet, schedule_df, makespan_val):
        makespan_hours = int(makespan_val / 60) + 2
        worksheet.cell(row=1, column=1, value="Makine \\ Saat")
        worksheet.column_dimensions['A'].width = 20
        
        for h in range(makespan_hours):
            col_letter = get_column_letter(h + 2)
            worksheet.cell(row=1, column=h+2, value=str(h))
            worksheet.column_dimensions[col_letter].width = 3
            
        macs = sorted(schedule_df['machine'].unique())
        r_idx = 2
        for m in macs:
            worksheet.cell(row=r_idx, column=1, value=m)
            m_tasks = schedule_df[schedule_df['machine'] == m]
            for _, task in m_tasks.iterrows():
                s_h = int(task['start_min'] / 60)
                e_h = int(task['end_min'] / 60)
                col_start = s_h + 2
                col_end = e_h + 2
                fill = PatternFill(start_color=get_color(task['unit_id']), fill_type="solid")
                for c_idx in range(col_start, col_end + 1):
                    worksheet.cell(row=r_idx, column=c_idx).fill = fill
                    if c_idx == col_start:
                        worksheet.cell(row=r_idx, column=c_idx).value = str(task['unit_id'])[:6]
            r_idx += 1
            
    draw_excel_gantt(ws, df_schedule, opt_gross)
    
    fifo_schedule_file = os.path.join("fifo_output", "schedule.csv")
    if os.path.exists(fifo_schedule_file):
        df_fifo = pd.read_csv(fifo_schedule_file)
        df_fifo.to_excel(writer, sheet_name="FIFO Çizelge (Liste)", index=False)
        ws_fifo = workbook.create_sheet("FIFO Gantt Görseli")
        draw_excel_gantt(ws_fifo, df_fifo, fifo_gross)

print("Excel generation successful. Size:", os.path.getsize(excel_path))
