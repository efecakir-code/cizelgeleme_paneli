import pandas as pd
import os
import hashlib
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = "custom_output"
excel_path = os.path.join(OUTPUT_DIR, "cizelge_sonuclari.xlsx")
schedule_file = os.path.join(OUTPUT_DIR, "schedule.csv")

df_schedule = pd.read_csv(schedule_file)
for col in df_schedule.select_dtypes(include=['datetimetz']).columns:
    df_schedule[col] = df_schedule[col].dt.tz_localize(None)

machine_file = os.path.join(OUTPUT_DIR, "machine_utilization.csv")
if os.path.exists(machine_file):
    df_machine = pd.read_csv(machine_file)
else:
    df_machine = pd.DataFrame({"Hata": ["Makine verisi bulunamadı."]})

df_obj = pd.DataFrame([{"Test": 1}])

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_obj.to_excel(writer, sheet_name="Amaç Fonksiyonları", index=False)
    df_machine.to_excel(writer, sheet_name="Makine Kullanımı", index=False)
    df_schedule.to_excel(writer, sheet_name="Çizelge (Liste)", index=False)
    
    workbook = writer.book
    ws = workbook.create_sheet("Opt. Gantt Görseli")
    
    def get_color(uid):
        hash_obj = hashlib.md5(str(uid).encode())
        hex_col = hash_obj.hexdigest()[:6]
        return "FF" + hex_col
        
    def draw_excel_gantt(worksheet, schedule_df, makespan_val):
        makespan_hours = int(makespan_val / 60) + 2
        worksheet.cell(row=1, column=1, value="Makine \ Saat")
        worksheet.column_dimensions['A'].width = 20
        
        for h in range(makespan_hours):
            col_letter = get_column_letter(h + 2)
            worksheet.cell(row=1, column=h+2, value=str(h))
            worksheet.column_dimensions[col_letter].width = 3
            
        macs = sorted(schedule_df['machine'].unique())
        r_idx = 2
        for m in macs:
            worksheet.cell(row=r_idx, column=1, value=m)
            m_tasks = schedule_df[schedule_df['machine'] == m]
            for _, task in m_tasks.iterrows():
                s_h = int(task['start_min'] / 60)
                e_h = int(task['end_min'] / 60)
                col_start = s_h + 2
                col_end = e_h + 2
                color = get_color(task['unit_id'])
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for c_idx in range(col_start, col_end + 1):
                    cell = worksheet.cell(row=r_idx, column=c_idx)
                    cell.fill = fill
                    if c_idx == col_start:
                        cell.value = str(task['unit_id']).replace("MOCK-", "").replace("NEW-", "")[:6]
            r_idx += 1
            
    draw_excel_gantt(ws, df_schedule, 500)
    print("Success")
