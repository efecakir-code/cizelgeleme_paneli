"""Güncel net imalat alternatörlerini onaylı rotalarla CP/GA girdisine dönüştür.

Kaynak Excel'de yalnızca ``Çizelgeye Al = Evet`` ve
``Yeni İmalat Adedi = 1`` olan fiziksel birimler alınır. Operasyon rotaları,
daha önce doğrulanmış ``master_three_stage_input.json`` dosyasındaki aynı
alternatör kodunun tekil rota şablonundan kopyalanır.
"""

from __future__ import annotations

import argparse
import copy
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


ROOT = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = ROOT / "157_alternator_stok_sonrasi_net_ihtiyac.xlsx"
DEFAULT_ROUTE_INPUT = ROOT / "data" / "master_three_stage_input.json"
DEFAULT_PRIORITY_WORKBOOK = ROOT / "FJSP_Endustri_Muhendisligi_Ana_Veri_Kesinlestirilmis(1).xlsx"
DEFAULT_OUTPUT = ROOT / "data" / "master_net_manufacturing_input.json"
DEFAULT_AUDIT = ROOT / "data" / "master_net_manufacturing_input_audit.json"
DAILY_WORK_MINUTES = 540


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def rows_as_dicts(sheet: Any) -> list[dict[str, Any]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    return [dict(zip(headers, values, strict=True)) for values in iterator if any(value is not None for value in values)]


def as_date(value: Any, epoch: datetime) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value, epoch)
        return converted.date() if isinstance(converted, datetime) else converted
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    raise ValueError(f"Tarih okunamadı: {value!r}")


def as_identifier(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def roll_forward_to_weekday(day: date) -> date:
    while day.weekday() >= 5:
        day += timedelta(days=1)
    return day


def business_minutes(day: date, base_date: date, scale: int) -> int:
    """Eski CP ile aynı: yalnız Pzt-Cuma günlerini 540 net dakika say."""

    if day < base_date:
        raise ValueError(f"{day}: baz tarihten ({base_date}) önce")
    workdays = sum(
        1
        for offset in range((day - base_date).days)
        if (base_date + timedelta(days=offset)).weekday() < 5
    )
    return workdays * DAILY_WORK_MINUTES * scale


def derive_priority(due_date: date, ready_date: date) -> str:
    if due_date <= ready_date:
        return "Acil"
    if (due_date - ready_date).days <= 30:
        return "Yüksek"
    return "Normal"


def priority_by_source(priority_workbook: Path | None) -> dict[int, str]:
    """Eski CP'nin Atomik Plan/Kaynak Satır önceliklerini oku."""

    if priority_workbook is None or not priority_workbook.exists():
        return {}
    workbook = load_workbook(priority_workbook, read_only=True, data_only=True)
    if "Atomik Plan" not in workbook.sheetnames:
        raise ValueError(f"{priority_workbook}: Atomik Plan sayfası yok")
    result: dict[int, str] = {}
    values = workbook["Atomik Plan"].iter_rows(values_only=True)
    headers: list[str] | None = None
    for raw in values:
        candidate = [str(value).strip() if value is not None else "" for value in raw]
        if "Kaynak Satır" in candidate and "Öncelik" in candidate:
            headers = candidate
            break
    if headers is None:
        raise ValueError(f"{priority_workbook}: Atomik Plan başlıkları bulunamadı")
    for raw in values:
        row = dict(zip(headers, raw, strict=True))
        source = row.get("Kaynak Satır")
        priority = str(row.get("Öncelik") or "").strip()
        if source not in (None, "") and priority in {"Acil", "Yüksek", "Normal"}:
            try:
                source_number = int(float(source))
            except (TypeError, ValueError):
                continue
            result.setdefault(source_number, priority)
    return result


def route_templates(route_data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    variants: dict[str, dict[str, str]] = defaultdict(dict)
    sample_by_code: dict[str, dict[str, Any]] = {}
    for unit in route_data["units"]:
        code = str(unit["alternator_code"])
        signature = json.dumps(unit["operations"], ensure_ascii=False, sort_keys=True)
        variants[code][signature] = unit["unit_id"]
        sample_by_code.setdefault(code, unit)
    inconsistent = {code: values for code, values in variants.items() if len(values) != 1}
    if inconsistent:
        raise ValueError(f"Bir alternatör kodunda birden fazla rota varyantı var: {sorted(inconsistent)}")
    return sample_by_code


def infer_base_date(workbook: Any) -> date:
    if "Tarihli_Net_Ihtiyac" in workbook.sheetnames:
        for cell in next(workbook["Tarihli_Net_Ihtiyac"].iter_rows(values_only=True)):
            match = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", str(cell or ""))
            if match:
                day, month, year = map(int, match.groups())
                return date(year, month, day)
    return date.today()


def build_payload(
    workbook_path: Path,
    route_input: Path,
    planning_start: date | None = None,
    priority_workbook: Path | None = DEFAULT_PRIORITY_WORKBOOK,
) -> tuple[dict[str, Any], dict[str, Any]]:
    route_data = json.loads(route_input.read_text(encoding="utf-8"))
    templates = route_templates(route_data)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    planning_start = roll_forward_to_weekday(planning_start or date.today())
    required_sheets = {"Net_Ihtiyac_Detay", "157_Tekil_Is"}
    if missing := required_sheets - set(workbook.sheetnames):
        raise ValueError(f"Kaynak Excel sayfası eksik: {sorted(missing)}")

    detail_rows = rows_as_dicts(workbook["Net_Ihtiyac_Detay"])
    unit_rows = rows_as_dicts(workbook["157_Tekil_Is"])
    unit_by_id = {str(row["Tekil İş ID"]).strip(): row for row in unit_rows}
    selected = [
        row
        for row in detail_rows
        if str(row.get("Çizelgeye Al") or "").strip().casefold() == "evet"
        and int(row.get("Yeni İmalat Adedi") or 0) == 1
    ]
    if not selected:
        raise ValueError("Excel'de çizelgelenecek yeni imalat satırı bulunamadı")

    scale = int(route_data["time_scale"])
    priority_weights = route_data["parameters"]["priority_weights"]
    source_priorities = priority_by_source(priority_workbook)
    units: list[dict[str, Any]] = []
    missing_routes: Counter[str] = Counter()
    code_counts: Counter[str] = Counter()
    priority_counts: Counter[str] = Counter()
    atomic_operations = 0

    prepared: list[tuple[dict[str, Any], dict[str, Any], date, date]] = []
    for detail in selected:
        unit_id = str(detail["Tekil İş ID"]).strip()
        source = unit_by_id.get(unit_id)
        if source is None:
            raise ValueError(f"{unit_id}: 157_Tekil_Is satırı bulunamadı")
        code = str(detail["Alternatör Stok Kodu"]).strip()
        template = templates.get(code)
        if template is None:
            missing_routes[code] += 1
            continue

        actual_start = as_date(source.get("Gerçekleşen Üretime Giriş"), workbook.epoch)
        planned_start = as_date(source.get("Planlanan Üretim Başlangıç Tarihleri"), workbook.epoch)
        ready_date = roll_forward_to_weekday(actual_start or planned_start or planning_start)
        ready_date = max(ready_date, planning_start)
        due_date = (
            as_date(source.get("Planlanan Sevk Tarihi"), workbook.epoch)
            or as_date(detail.get("Planlanan Sevk Tarihi"), workbook.epoch)
            or as_date(source.get("Teslim Tarihi"), workbook.epoch)
            or as_date(detail.get("Teslim Tarihi"), workbook.epoch)
        )
        if due_date is None:
            raise ValueError(f"{unit_id}: termin tarihi bulunamadı")
        due_date = roll_forward_to_weekday(due_date)
        prepared.append((detail, source, ready_date, due_date))

    if missing_routes:
        raise ValueError(f"Rota şablonu bulunmayan kodlar: {dict(missing_routes)}")
    base_date = roll_forward_to_weekday(
        min(day for _, _, ready, due in prepared for day in (ready, due))
    )

    for detail, source, ready_date, due_date in prepared:
        unit_id = str(detail["Tekil İş ID"]).strip()
        code = str(detail["Alternatör Stok Kodu"]).strip()
        template = templates[code]
        source_row = int(detail["Kaynak Excel Satırı"])
        priority = source_priorities.get(source_row) or derive_priority(due_date, ready_date)
        release = business_minutes(ready_date, base_date, scale)
        # Eski CP'de termin, termin günü vardiya sonudur.
        due = business_minutes(due_date, base_date, scale) + DAILY_WORK_MINUTES * scale
        operations = copy.deepcopy(template["operations"])
        atomic_operations += len(operations)
        code_counts[code] += 1
        priority_counts[priority] += 1
        units.append(
            {
                "unit_id": unit_id,
                "order_job_id": as_identifier(detail["Canias Sipariş No"]),
                "alternator_code": code,
                "model": template["model"],
                "series": template["series"],
                "priority": priority,
                "priority_weight": int(priority_weights[priority]),
                "release": release,
                "due": due,
                "source_excel_row": source_row,
                "ready_date": ready_date.isoformat(),
                "planned_production_start": as_date(
                    source.get("Planlanan Üretim Başlangıç Tarihleri"), workbook.epoch
                ).isoformat()
                if source.get("Planlanan Üretim Başlangıç Tarihleri") not in (None, "")
                else None,
                "planned_ship_date": due_date.isoformat(),
                "model_due_date": due_date.isoformat(),
                "project": str(detail.get("Proje Detayı") or "").strip(),
                "stock_confirmation_status": str(detail.get("Stok Teyit Durumu") or "").strip(),
                "operations": operations,
            }
        )

    if len({unit["unit_id"] for unit in units}) != len(units):
        raise ValueError("Tekil iş kimlikleri benzersiz değil")

    payload = {
        "schema_version": 2,
        "model": route_data["model"],
        "scenario_mode": "net_manufacturing_dynamic",
        "unit_count": len(units),
        "base_date": datetime.combine(base_date, datetime.min.time()).isoformat(),
        "source_workbook": str(workbook_path.resolve()),
        "source_sha256": sha256_file(workbook_path),
        "route_template_input": str(route_input.resolve()),
        "route_template_sha256": sha256_file(route_input),
        "time_unit": route_data.get("time_unit", "minute"),
        "time_scale": scale,
        "time_axis": "net_work_minutes",
        "daily_work_minutes": DAILY_WORK_MINUTES,
        "planning_start": planning_start.isoformat(),
        "parameters": copy.deepcopy(route_data["parameters"]),
        "machines": list(route_data["machines"]),
        "units": units,
    }
    audit = {
        "status": "PASS",
        "source_workbook": str(workbook_path.resolve()),
        "base_date": base_date.isoformat(),
        "planning_start": planning_start.isoformat(),
        "daily_work_minutes": DAILY_WORK_MINUTES,
        "calendar_rule": "Yalnız Pazartesi-Cuma; gece, öğle arası ve hafta sonu Cmax'a eklenmez",
        "gross_physical_units": len(detail_rows),
        "scheduled_new_manufacturing_units": len(units),
        "excluded_stock_or_in_production_units": len(detail_rows) - len(units),
        "atomic_operations": atomic_operations,
        "machines": len(payload["machines"]),
        "alternator_code_counts": dict(sorted(code_counts.items())),
        "priority_rule": "Eski CP Atomik Plan/Kaynak Satır; eksikse termin-ready farkından türetim",
        "priority_workbook": str(priority_workbook.resolve()) if priority_workbook and priority_workbook.exists() else None,
        "priority_counts": dict(priority_counts),
        "route_variant_check": "Her alternatör kodu için tek onaylı rota varyantı",
    }
    return payload, audit


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--route-input", type=Path, default=DEFAULT_ROUTE_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    parser.add_argument(
        "--planning-start",
        type=date.fromisoformat,
        default=None,
        help="Yeniden planlamanın ilk iş günü; boşsa bugün.",
    )
    parser.add_argument("--priority-workbook", type=Path, default=DEFAULT_PRIORITY_WORKBOOK)
    parser.add_argument(
        "--expected-units",
        type=int,
        default=None,
        help="İsteğe bağlı emniyet kontrolü; güncel çalışma için 143 verilebilir.",
    )
    args = parser.parse_args()
    payload, audit = build_payload(
        args.workbook,
        args.route_input,
        args.planning_start,
        args.priority_workbook,
    )
    if args.expected_units is not None and len(payload["units"]) != args.expected_units:
        raise ValueError(
            f"Beklenen {args.expected_units}, Excel'den seçilen {len(payload['units'])} yeni imalat var"
        )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    args.audit.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(audit, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
